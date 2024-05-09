import openpyxl
from PIL import Image

# Load the image
img = Image.open('download_image.png')

# Resize the image to a smaller size for better performance
img.thumbnail((img.width, img.height))

# Convert the image to RGB
rgb_img = img.convert('RGB')

# Create a new workbook and get the active sheet
wb = openpyxl.Workbook()
ws = wb.active

# Iterate over the pixels in the image
for i in range(rgb_img.width):
    for j in range(rgb_img.height):
        # Get the RGB values of the pixel
        r, g, b = rgb_img.getpixel((i, j))

        # Convert the RGB values to a hex color code
        color = "{:02x}{:02x}{:02x}".format(r, g, b)

        # Set the background color of the cell at the corresponding position to the color of the pixel
        ws.cell(row=j + 1, column=i + 1).fill = openpyxl.styles.PatternFill(start_color=color, end_color=color,
                                                                            fill_type="solid")

# Save the workbook
wb.save('output.xlsx')

# Print a success message
print("The image was successfully drawn in the Excel file 'output.xlsx'.")